import streamlit as st
import pandas as pd
from io import BytesIO
from datetime import datetime
from zoneinfo import ZoneInfo
import uuid
from urllib.parse import quote
from xml.sax.saxutils import escape
import unicodedata

from openpyxl.styles import Alignment, Font, PatternFill
from reportlab.platypus import (
    SimpleDocTemplate,
    Table,
    TableStyle,
    Paragraph,
    Spacer
)
from reportlab.lib import colors
from reportlab.lib.pagesizes import A4, landscape
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.enums import TA_CENTER, TA_LEFT

TZ = ZoneInfo("America/Bahia")


def agora_brasil():
    return datetime.now(TZ)


def hoje_brasil():
    return agora_brasil().date()


def normalizar_whatsapp(numero):
    if numero is None:
        return None

    numero = "".join(
        caractere
        for caractere in str(numero)
        if caractere.isdigit()
    )

    if not numero:
        return None

    if numero.startswith("55"):
        if len(numero) in (12, 13):
            return numero

        return None

    if len(numero) in (10, 11):
        return "55" + numero

    return None


def carregar_config(supabase):
    resp = (
        supabase
        .table("config_permissionarios")
        .select("*")
        .eq("id", 1)
        .limit(1)
        .execute()
    )

    dados = resp.data or []

    mensagem_padrao = (
        "Obrigada, {nome}! Recebemos suas informações com sucesso. "
        "Agradecemos sua colaboração com a cotação diária."
    )

    if dados:
        config = dados[0]
        config["mensagem_agradecimento"] = str(
            config.get("mensagem_agradecimento") or mensagem_padrao
        )
        return config

    return {
        "id": 1,
        "hora_envio": "07:00:00",
        "hora_limite": "09:00:00",
        "mensagem": (
            "Bom dia! Por favor, informe os preços dos produtos "
            "solicitados para a cotação de hoje."
        ),
        "mensagem_agradecimento": mensagem_padrao,
        "base_url": "",
        "ativo": False,
        "template_nome": "link_cotacao_diaria",
        "idioma": "pt_BR"
    }


def atualizar_config(
    supabase,
    hora_envio,
    hora_limite,
    mensagem,
    mensagem_agradecimento,
    base_url,
    ativo,
    template_nome,
    idioma
):
    supabase.table("config_permissionarios").upsert({
        "id": 1,
        "hora_envio": str(hora_envio),
        "hora_limite": str(hora_limite),
        "mensagem": mensagem.strip(),
        "mensagem_agradecimento": mensagem_agradecimento.strip(),
        "base_url": base_url.strip().rstrip("/"),
        "ativo": bool(ativo),
        "template_nome": template_nome.strip(),
        "idioma": idioma.strip(),
        "atualizado_em": agora_brasil().isoformat()
    }).execute()


def carregar_permissionarios(supabase):
    resp = supabase.table("permissionarios").select("*").order("nome").execute()
    return pd.DataFrame(resp.data or [])

def carregar_disparos_do_dia(supabase):
    data_cotacao = hoje_brasil().isoformat()

    resposta = (
        supabase
        .table("disparos_whatsapp")
        .select(
            "permissionario_id, status, "
            "message_id, enviado_em"
        )
        .eq("data_cotacao", data_cotacao)
        .execute()
    )

    return {
        int(item["permissionario_id"]): item
        for item in (resposta.data or [])
    }

def carregar_vinculos(supabase, permissionario_id):
    resp = supabase.table("permissionario_produtos")\
        .select("*")\
        .eq("permissionario_id", int(permissionario_id))\
        .eq("ativo", True)\
        .order("produto")\
        .execute()

    return pd.DataFrame(resp.data or [])


def salvar_vinculos(supabase, permissionario_id, produtos_df, produtos_selecionados):
    supabase.table("permissionario_produtos")\
        .delete()\
        .eq("permissionario_id", int(permissionario_id))\
        .execute()

    dados = []

    for produto in produtos_selecionados:
        linha = produtos_df[produtos_df["nome"] == produto]

        if not linha.empty:
            classe = str(linha.iloc[0].get("classe", ""))
        else:
            classe = ""

        dados.append({
            "permissionario_id": int(permissionario_id),
            "produto": str(produto).strip().upper(),
            "classe": classe,
            "ativo": True
        })

    if dados:
        supabase.table("permissionario_produtos").insert(dados).execute()



def montar_validade_link(config):
    hora_limite_txt = str(
        config.get("hora_limite") or "09:00:00"
    )[:5]

    hora_limite = datetime.strptime(
        hora_limite_txt,
        "%H:%M"
    ).time()

    return datetime.combine(
        hoje_brasil(),
        hora_limite,
        tzinfo=TZ
    )


def gerar_ou_atualizar_link(supabase, permissionario_id, config):
    data_cotacao = hoje_brasil().isoformat()
    valido_ate = montar_validade_link(config)

    base_url = str(
        config.get("base_url") or ""
    ).strip().rstrip("/")

    if not base_url:
        raise ValueError(
            "A URL do sistema publicado não foi configurada."
        )

    resposta = (
        supabase
        .table("links_permissionarios")
        .select("*")
        .eq("permissionario_id", int(permissionario_id))
        .eq("data", data_cotacao)
        .limit(1)
        .execute()
    )

    if resposta.data:
        registro = resposta.data[0]
        token = registro["token"]

        # Um link já utilizado não pode ser reativado.
        if not bool(registro.get("usado", False)):
            (
                supabase
                .table("links_permissionarios")
                .update({
                    "valido_ate": valido_ate.isoformat()
                })
                .eq("id", registro["id"])
                .execute()
            )
    else:
        token = uuid.uuid4().hex

        (
            supabase
            .table("links_permissionarios")
            .insert({
                "permissionario_id": int(permissionario_id),
                "data": data_cotacao,
                "token": token,
                "valido_ate": valido_ate.isoformat(),
                "usado": False
            })
            .execute()
        )

    return f"{base_url}?token={token}", valido_ate


def registrar_disparo(
    supabase,
    data_cotacao,
    permissionario_id,
    permissionario_nome,
    whatsapp,
    link,
    status,
    message_id=None,
    erro=None
):
    dados = {
        "data_cotacao": data_cotacao,
        "permissionario_id": int(permissionario_id),
        "permissionario_nome": permissionario_nome,
        "whatsapp": whatsapp,
        "link": link,
        "status": status,
        "message_id": message_id,
        "erro": erro,
        "enviado_em": agora_brasil().isoformat()
    }

    resposta = (
        supabase
        .table("disparos_whatsapp")
        .select("id")
        .eq("data_cotacao", data_cotacao)
        .eq("permissionario_id", int(permissionario_id))
        .limit(1)
        .execute()
    )

    if resposta.data:
        (
            supabase
            .table("disparos_whatsapp")
            .update(dados)
            .eq("id", resposta.data[0]["id"])
            .execute()
        )
    else:
        (
            supabase
            .table("disparos_whatsapp")
            .insert(dados)
            .execute()
        )


def excluir_permissionario_completo(supabase, permissionario_id):
    permissionario_id = int(permissionario_id)

    # Tenta remover também os arquivos de fotos do Storage.
    try:
        resposta_fotos = (
            supabase
            .table("fotos_permissionarios")
            .select("arquivo_nome")
            .eq("permissionario_id", permissionario_id)
            .execute()
        )

        arquivos = [
            item.get("arquivo_nome")
            for item in (resposta_fotos.data or [])
            if item.get("arquivo_nome")
        ]

        if arquivos:
            supabase.storage.from_(BUCKET_FOTOS).remove(arquivos)
    except Exception:
        # A exclusão do cadastro não será interrompida se a limpeza
        # de algum arquivo antigo do Storage falhar.
        pass

    tabelas_relacionadas = [
        "respostas_permissionarios",
        "fotos_permissionarios",
        "permissionario_produtos",
        "links_permissionarios",
        "disparos_whatsapp"
    ]

    for tabela in tabelas_relacionadas:
        (
            supabase
            .table(tabela)
            .delete()
            .eq("permissionario_id", permissionario_id)
            .execute()
        )

    (
        supabase
        .table("permissionarios")
        .delete()
        .eq("id", permissionario_id)
        .execute()
    )


def montar_mensagem_whatsapp(nome, link, config):
    mensagem = str(config.get("mensagem") or "").strip()

    if mensagem == "":
        mensagem = (
            "Bom dia! Por favor, informe os preços dos produtos "
            "solicitados para a cotação de hoje."
        )

    saudacao = f"Olá, {nome}!" if nome else "Olá!"

    return (
        f"{saudacao}\n\n"
        f"{mensagem}\n\n"
        f"Link para preencher:\n{link}"
    )


def carregar_respostas_permissionarios(supabase, data_str):
    try:
        resp = supabase.table("respostas_permissionarios")\
            .select("*")\
            .eq("data", data_str)\
            .order("produto")\
            .execute()

        df = pd.DataFrame(resp.data or [])

        if not df.empty:
            df["produto"] = df["produto"].astype(str).str.strip().str.upper()
            df["preco"] = pd.to_numeric(df["preco"], errors="coerce")

        return df

    except Exception as e:
        st.error(f"Erro ao carregar respostas dos permissionários: {e}")
        return pd.DataFrame()
    
BUCKET_FOTOS = "fotos-produtos"

def limpar_nome_arquivo(texto):
    texto = str(texto).lower().strip()

    # remove acentos
    texto = unicodedata.normalize("NFKD", texto)
    texto = texto.encode("ascii", "ignore").decode("ascii")

    # troca caracteres especiais por _
    texto = "".join(
        c if c.isalnum() else "_"
        for c in texto
    )

    # remove underscores repetidos
    while "__" in texto:
        texto = texto.replace("__", "_")

    # remove _ do começo e do final
    texto = texto.strip("_")

    if texto == "":
        texto = "produto"

    return texto[:60]

def salvar_foto_permissionario(supabase, arquivo, permissionario_id, data_link, produto):
    nome_original = str(arquivo.name)
    extensao = nome_original.split(".")[-1].lower()

    if extensao not in ["jpg", "jpeg", "png", "webp"]:
        extensao = "jpg"

    produto_limpo = limpar_nome_arquivo(produto)

    nome_arquivo = (
        f"permissionarios/{data_link}/{permissionario_id}/"
        f"{produto_limpo}_{uuid.uuid4().hex}.{extensao}"
    )

    content_type = arquivo.type or f"image/{extensao}"

    supabase.storage.from_(BUCKET_FOTOS).upload(
        nome_arquivo,
        arquivo.getvalue(),
        {
            "content-type": content_type,
            "upsert": "true"
        }
    )

    url_foto = supabase.storage.from_(BUCKET_FOTOS).get_public_url(nome_arquivo)

    return url_foto, nome_arquivo


def carregar_fotos_permissionarios(supabase, data_str):
    try:
        resp = supabase.table("fotos_permissionarios")\
            .select("*")\
            .eq("data", data_str)\
            .order("produto")\
            .execute()

        df = pd.DataFrame(resp.data or [])

        if not df.empty:
            df["produto"] = df["produto"].astype(str).str.strip().str.upper()

        return df

    except Exception as e:
        st.error(f"Erro ao carregar fotos dos permissionários: {e}")
        return pd.DataFrame()


def converter_preco_digitado(valor):
    texto = str(valor or "").strip()

    if not texto:
        return None

    texto = (
        texto
        .replace("R$", "")
        .replace(" ", "")
    )

    if "," in texto and "." in texto:
        if texto.rfind(",") > texto.rfind("."):
            texto = texto.replace(".", "").replace(",", ".")
        else:
            texto = texto.replace(",", "")
    else:
        texto = texto.replace(",", ".")

    try:
        preco = float(texto)
    except (TypeError, ValueError):
        return None

    if preco <= 0:
        return None

    return preco


def mostrar_agradecimento(config, nome, mostrar_confetes=False):
    mensagem_padrao = (
        "Obrigada, {nome}! Recebemos suas informações com sucesso. "
        "Agradecemos sua colaboração com a cotação diária."
    )

    mensagem = str(
        config.get("mensagem_agradecimento")
        or mensagem_padrao
    ).strip()

    mensagem = mensagem.replace("{nome}", str(nome or "").strip())

    if mostrar_confetes:
        st.balloons()

    st.title("✅ Envio concluído")
    st.success(mensagem)
    st.info(
        "As informações foram registradas. "
        "Este link já foi utilizado e não permite um novo envio."
    )


def tela_publica_permissionario(supabase, token):
    try:
        resp_link = (
            supabase
            .table("links_permissionarios")
            .select("*")
            .eq("token", token)
            .limit(1)
            .execute()
        )

        dados_link = resp_link.data or []

        if not dados_link:
            st.title("🧾 Envio de Preços")
            st.error("Link inválido.")
            return

        link = dados_link[0]
        permissionario_id = int(link["permissionario_id"])
        data_link = str(link["data"])

        resp_perm = (
            supabase
            .table("permissionarios")
            .select("*")
            .eq("id", permissionario_id)
            .limit(1)
            .execute()
        )

        dados_perm = resp_perm.data or []

        if not dados_perm:
            st.title("🧾 Envio de Preços")
            st.error("Permissionário não encontrado.")
            return

        permissionario = dados_perm[0]
        nome_permissionario = str(
            permissionario.get("nome") or ""
        ).strip()

        config = carregar_config(supabase)
        chave_confetes = f"confetes_permissionario_{token}"

        # Se já houve envio, o formulário não aparece novamente.
        if bool(link.get("usado", False)):
            mostrar_confetes = bool(
                st.session_state.pop(chave_confetes, False)
            )

            mostrar_agradecimento(
                config,
                nome_permissionario,
                mostrar_confetes=mostrar_confetes
            )
            return

        valido_ate = pd.to_datetime(link["valido_ate"])
        agora = agora_brasil()

        if valido_ate.tzinfo is None:
            valido_ate = valido_ate.tz_localize(
                "America/Bahia"
            )

        if agora > valido_ate.to_pydatetime():
            st.title("🧾 Envio de Preços")
            st.error(
                "Este link expirou. Entre em contato com a administração."
            )
            return

        if not permissionario.get("ativo", True):
            st.title("🧾 Envio de Preços")
            st.error("Permissionário inativo.")
            return

        st.title("🧾 Envio de Preços")
        st.markdown(
            f"### Bom dia, {nome_permissionario}!"
        )
        st.info(
            "Informe os preços dos produtos solicitados. "
            f"Link válido até {valido_ate.strftime('%H:%M')}."
        )

        df_produtos = carregar_vinculos(
            supabase,
            permissionario_id
        )

        if df_produtos.empty:
            st.warning(
                "Nenhum produto vinculado para este permissionário."
            )
            return

        with st.form("form_resposta_permissionario"):
            respostas = []
            fotos_para_salvar = []

            qtd_precos = int(
                permissionario.get("qtd_precos", 1) or 1
            )

            for _, row in df_produtos.iterrows():
                produto = str(
                    row["produto"]
                ).strip().upper()

                classe = str(row.get("classe", ""))

                st.markdown(f"### {produto}")

                cols = st.columns(3)

                for i in range(qtd_precos):
                    with cols[i % 3]:
                        preco_texto = st.text_input(
                            f"Preço {i + 1}",
                            value="",
                            placeholder="Digite o preço",
                            key=(
                                f"preco_publico_"
                                f"{permissionario_id}_"
                                f"{produto}_{i}"
                            )
                        )

                        respostas.append({
                            "permissionario_id": permissionario_id,
                            "permissionario_nome": nome_permissionario,
                            "data": data_link,
                            "produto": produto,
                            "classe": classe,
                            "numero_preco": i + 1,
                            "preco_texto": preco_texto,
                            "preco": converter_preco_digitado(
                                preco_texto
                            )
                        })

                fotos = st.file_uploader(
                    f"Fotos de {produto}",
                    type=["jpg", "jpeg", "png", "webp"],
                    accept_multiple_files=True,
                    key=(
                        f"fotos_publico_"
                        f"{permissionario_id}_{produto}"
                    )
                )

                if fotos:
                    for foto in fotos:
                        fotos_para_salvar.append({
                            "produto": produto,
                            "classe": classe,
                            "arquivo": foto
                        })

                st.divider()

            enviar = st.form_submit_button(
                "Enviar preços e fotos",
                type="primary"
            )

            if enviar:
                campos_invalidos = [
                    resposta
                    for resposta in respostas
                    if (
                        str(
                            resposta.get("preco_texto") or ""
                        ).strip()
                        and resposta.get("preco") is None
                    )
                ]

                if campos_invalidos:
                    st.warning(
                        "Há um preço inválido. Use apenas números, "
                        "por exemplo: 12,50."
                    )
                    return

                dados_validos = []

                for resposta in respostas:
                    if resposta.get("preco") is not None:
                        dados_validos.append({
                            chave: valor
                            for chave, valor in resposta.items()
                            if chave != "preco_texto"
                        })

                if not dados_validos and not fotos_para_salvar:
                    st.warning(
                        "Preencha pelo menos um preço ou envie uma foto "
                        "antes de concluir."
                    )
                    return

                (
                    supabase
                    .table("respostas_permissionarios")
                    .delete()
                    .eq("permissionario_id", permissionario_id)
                    .eq("data", data_link)
                    .execute()
                )

                if dados_validos:
                    (
                        supabase
                        .table("respostas_permissionarios")
                        .insert(dados_validos)
                        .execute()
                    )

                if fotos_para_salvar:
                    (
                        supabase
                        .table("fotos_permissionarios")
                        .delete()
                        .eq("permissionario_id", permissionario_id)
                        .eq("data", data_link)
                        .execute()
                    )

                    dados_fotos = []

                    for item in fotos_para_salvar:
                        url_foto, nome_arquivo = salvar_foto_permissionario(
                            supabase,
                            item["arquivo"],
                            permissionario_id,
                            data_link,
                            item["produto"]
                        )

                        dados_fotos.append({
                            "permissionario_id": permissionario_id,
                            "permissionario_nome": nome_permissionario,
                            "data": data_link,
                            "produto": item["produto"],
                            "classe": item["classe"],
                            "foto_url": url_foto,
                            "arquivo_nome": nome_arquivo
                        })

                    if dados_fotos:
                        (
                            supabase
                            .table("fotos_permissionarios")
                            .insert(dados_fotos)
                            .execute()
                        )

                (
                    supabase
                    .table("links_permissionarios")
                    .update({"usado": True})
                    .eq("token", token)
                    .execute()
                )

                st.session_state[chave_confetes] = True
                st.rerun()

    except Exception as erro:
        st.title("🧾 Envio de Preços")
        st.error(f"Erro ao abrir link: {erro}")


def aba_cadastro_permissionario(supabase, registrar_acao):
    st.subheader("➕ Cadastrar permissionário")

    col_nome, col_whatsapp = st.columns(2)

    with col_nome:
        nome = st.text_input(
            "Nome do permissionário",
            key="perm_nome"
        )

    with col_whatsapp:
        whatsapp_digitado = st.text_input(
            "WhatsApp",
            placeholder="Ex.: 74999999999",
            key="perm_whatsapp"
        )

    qtd_precos = st.number_input(
        "Quantidade de preços por produto",
        min_value=1,
        max_value=10,
        value=1,
        step=1,
        key="perm_qtd_precos"
    )

    if st.button(
        "Cadastrar permissionário",
        type="primary",
        key="btn_cadastrar_permissionario"
    ):
        nome_limpo = nome.strip()
        whatsapp = normalizar_whatsapp(whatsapp_digitado)

        if not nome_limpo:
            st.warning("Informe o nome do permissionário.")
            return

        if not whatsapp:
            st.warning(
                "Informe um WhatsApp válido com DDD."
            )
            return

        try:
            (
                supabase
                .table("permissionarios")
                .insert({
                    "nome": nome_limpo,
                    "whatsapp": whatsapp,
                    "qtd_precos": int(qtd_precos),
                    "ativo": True
                })
                .execute()
            )

            registrar_acao(
                "Cadastro de permissionário",
                "Permissionários",
                f"Permissionário cadastrado: {nome_limpo}"
            )

            st.success("Permissionário cadastrado com sucesso.")
            st.rerun()

        except Exception as erro:
            st.error(
                f"Erro ao cadastrar permissionário: {erro}"
            )

    st.divider()
    st.subheader("📋 Permissionários cadastrados")

    df_perm = carregar_permissionarios(supabase)

    if df_perm.empty:
        st.info("Nenhum permissionário cadastrado.")
        return

    colunas = [
        coluna
        for coluna in [
            "id",
            "nome",
            "whatsapp",
            "qtd_precos",
            "ativo"
        ]
        if coluna in df_perm.columns
    ]

    st.dataframe(
        df_perm[colunas],
        use_container_width=True,
        hide_index=True
    )


def aba_editar_excluir_permissionario(supabase, registrar_acao):
    st.subheader("✏️ Editar ou excluir permissionário")

    df_perm = carregar_permissionarios(supabase)

    if df_perm.empty:
        st.info("Nenhum permissionário cadastrado.")
        return

    ids = df_perm["id"].astype(int).tolist()
    nomes_por_id = {
        int(linha["id"]): str(linha.get("nome", ""))
        for _, linha in df_perm.iterrows()
    }

    id_selecionado = st.selectbox(
        "Selecione o permissionário",
        ids,
        format_func=lambda valor: (
            f"{nomes_por_id.get(int(valor), '')} — ID {valor}"
        ),
        key="perm_editar_id"
    )

    dados = (
        df_perm[df_perm["id"].astype(int) == int(id_selecionado)]
        .iloc[0]
    )

    st.caption(
        "Para manter o histórico, prefira inativar. "
        "A exclusão apaga também vínculos, links, respostas, "
        "fotos registradas e controles de envio."
    )

    col_nome, col_whatsapp = st.columns(2)

    with col_nome:
        novo_nome = st.text_input(
            "Nome",
            value=str(dados.get("nome", "")),
            key=f"edit_perm_nome_{id_selecionado}"
        )

    with col_whatsapp:
        novo_whatsapp = st.text_input(
            "WhatsApp",
            value=str(dados.get("whatsapp", "")),
            key=f"edit_perm_whatsapp_{id_selecionado}"
        )

    col_qtd, col_ativo = st.columns(2)

    with col_qtd:
        qtd_precos = st.number_input(
            "Quantidade de preços por produto",
            min_value=1,
            max_value=10,
            value=int(dados.get("qtd_precos", 1) or 1),
            step=1,
            key=f"edit_perm_qtd_{id_selecionado}"
        )

    with col_ativo:
        ativo = st.checkbox(
            "Permissionário ativo",
            value=bool(dados.get("ativo", True)),
            key=f"edit_perm_ativo_{id_selecionado}"
        )

    if st.button(
        "Salvar alterações",
        type="primary",
        key=f"btn_atualizar_perm_{id_selecionado}"
    ):
        nome_limpo = novo_nome.strip()
        whatsapp = normalizar_whatsapp(novo_whatsapp)

        if not nome_limpo:
            st.warning("Informe o nome do permissionário.")
        elif not whatsapp:
            st.warning(
                "Informe um WhatsApp válido com DDD."
            )
        else:
            try:
                (
                    supabase
                    .table("permissionarios")
                    .update({
                        "nome": nome_limpo,
                        "whatsapp": whatsapp,
                        "qtd_precos": int(qtd_precos),
                        "ativo": bool(ativo)
                    })
                    .eq("id", int(id_selecionado))
                    .execute()
                )

                registrar_acao(
                    "Atualização de permissionário",
                    "Permissionários",
                    f"Permissionário atualizado: {nome_limpo}"
                )

                st.success("Permissionário atualizado com sucesso.")
                st.rerun()

            except Exception as erro:
                st.error(
                    f"Erro ao atualizar permissionário: {erro}"
                )

    st.divider()
    st.subheader("🗑️ Excluir definitivamente")

    st.warning(
        "Essa ação não pode ser desfeita. Todos os dados ligados "
        "a esse permissionário serão apagados."
    )

    confirmar = st.checkbox(
        "Confirmo que desejo excluir este permissionário e seus dados.",
        key=f"confirmar_exclusao_perm_{id_selecionado}"
    )

    texto_confirmacao = st.text_input(
        'Digite EXCLUIR para confirmar',
        key=f"texto_exclusao_perm_{id_selecionado}"
    )

    pode_excluir = (
        confirmar
        and texto_confirmacao.strip().upper() == "EXCLUIR"
    )

    if st.button(
        "Excluir permissionário",
        disabled=not pode_excluir,
        key=f"btn_excluir_perm_{id_selecionado}"
    ):
        nome_excluido = str(dados.get("nome", ""))

        try:
            excluir_permissionario_completo(
                supabase,
                id_selecionado
            )

            registrar_acao(
                "Exclusão de permissionário",
                "Permissionários",
                f"Permissionário excluído: {nome_excluido}"
            )

            st.success("Permissionário excluído com sucesso.")
            st.rerun()

        except Exception as erro:
            st.error(
                "Não foi possível excluir o permissionário. "
                f"Detalhes: {erro}"
            )


def aba_produtos_permissionario(
    supabase,
    carregar_produtos,
    corrigir_classe,
    registrar_acao
):
    st.subheader("📦 Vincular produtos ao permissionário")

    df_perm = carregar_permissionarios(supabase)

    if df_perm.empty:
        st.info("Cadastre um permissionário primeiro.")
        return

    df_perm = df_perm.copy()
    df_perm["id"] = pd.to_numeric(
        df_perm["id"],
        errors="coerce"
    )
    df_perm = df_perm.dropna(subset=["id"])
    df_perm["id"] = df_perm["id"].astype(int)

    ids = df_perm["id"].tolist()
    nomes_por_id = {
        int(linha["id"]): str(linha.get("nome", ""))
        for _, linha in df_perm.iterrows()
    }

    permissionario_id = st.selectbox(
        "Permissionário",
        ids,
        format_func=lambda valor: nomes_por_id.get(int(valor), ""),
        key="perm_prod_id"
    )

    nome_selecionado = nomes_por_id.get(
        int(permissionario_id),
        ""
    )

    produtos = carregar_produtos()

    if produtos.empty:
        st.warning("Nenhum produto cadastrado no sistema.")
        return

    produtos = produtos.copy()
    produtos["nome"] = (
        produtos["nome"]
        .astype(str)
        .str.strip()
        .str.upper()
    )
    produtos["classe"] = produtos["classe"].apply(
        corrigir_classe
    )
    produtos = produtos.sort_values(["classe", "nome"])

    vinculos = carregar_vinculos(
        supabase,
        permissionario_id
    )

    produtos_atuais = []

    if not vinculos.empty:
        produtos_atuais = sorted(
            vinculos["produto"]
            .astype(str)
            .str.strip()
            .str.upper()
            .dropna()
            .unique()
            .tolist()
        )

    if produtos_atuais:
        st.success(
            f"{nome_selecionado} possui "
            f"{len(produtos_atuais)} produto(s) vinculado(s)."
        )
        st.caption("Produtos atuais: " + ", ".join(produtos_atuais))
    else:
        st.info(
            f"{nome_selecionado} ainda não possui produtos vinculados."
        )

    lista_produtos = produtos["nome"].drop_duplicates().tolist()

    chave_produtos = (
        f"multi_prod_permissionario_{permissionario_id}"
    )

    selecionados = st.multiselect(
        "Produtos",
        lista_produtos,
        default=[
            produto
            for produto in produtos_atuais
            if produto in lista_produtos
        ],
        key=chave_produtos
    )

    if st.button(
        "Salvar produtos vinculados",
        type="primary",
        key=f"btn_salvar_produtos_permissionario_{permissionario_id}"
    ):
        try:
            salvar_vinculos(
                supabase,
                permissionario_id,
                produtos,
                selecionados
            )

            registrar_acao(
                "Vínculo de produtos",
                "Permissionários",
                (
                    f"Produtos vinculados para {nome_selecionado}: "
                    f"{len(selecionados)}"
                )
            )

            st.success("Produtos vinculados com sucesso.")
            st.session_state.pop(chave_produtos, None)
            st.rerun()

        except Exception as erro:
            st.error(
                f"Erro ao salvar vínculos: {erro}"
            )



def aba_mensagem_link(supabase, registrar_acao):
    st.subheader("✉️ Mensagem e validade do link")

    config = carregar_config(supabase)

    hora_limite_txt = str(
        config.get("hora_limite") or "09:00"
    )[:5]

    mensagem = st.text_area(
        "Mensagem que será aberta no WhatsApp",
        value=str(config.get("mensagem") or ""),
        height=130,
        key="cfg_mensagem_perm"
    )

    mensagem_agradecimento = st.text_area(
        "Mensagem de agradecimento após o envio",
        value=str(
            config.get("mensagem_agradecimento") or ""
        ),
        height=130,
        help=(
            "Use {nome} para inserir automaticamente o nome "
            "do permissionário."
        ),
        key="cfg_mensagem_agradecimento"
    )

    base_url = st.text_input(
        "URL do sistema publicado",
        value=str(config.get("base_url") or ""),
        help=(
            "Use o endereço público completo do sistema, sem barra "
            "no final. Exemplo: https://seu-sistema.onrender.com"
        ),
        key="cfg_base_url"
    )

    hora_limite = st.time_input(
        "Horário limite para responder",
        value=datetime.strptime(
            hora_limite_txt,
            "%H:%M"
        ).time(),
        key="cfg_hora_limite"
    )

    st.info(
        "O envio é manual e gratuito. O sistema prepara uma "
        "mensagem e um link individual para cada permissionário."
    )

    if st.button(
        "Salvar mensagem e link",
        type="primary",
        key="btn_salvar_config_manual"
    ):
        base_url_limpa = base_url.strip().rstrip("/")

        if not mensagem.strip():
            st.error("Informe a mensagem do WhatsApp.")
        elif not mensagem_agradecimento.strip():
            st.error("Informe a mensagem de agradecimento.")
        elif not base_url_limpa:
            st.error("Informe a URL do sistema publicado.")
        elif not base_url_limpa.startswith(("http://", "https://")):
            st.error(
                "A URL deve começar com http:// ou https://."
            )
        else:
            try:
                atualizar_config(
                    supabase=supabase,
                    hora_envio=str(
                        config.get("hora_envio") or "07:00:00"
                    ),
                    hora_limite=hora_limite,
                    mensagem=mensagem,
                    mensagem_agradecimento=mensagem_agradecimento,
                    base_url=base_url_limpa,
                    ativo=False,
                    template_nome=str(
                        config.get("template_nome")
                        or "link_cotacao_diaria"
                    ),
                    idioma=str(
                        config.get("idioma") or "pt_BR"
                    )
                )

                registrar_acao(
                    "Configuração de mensagens",
                    "Permissionários",
                    (
                        "Envio manual configurado | "
                        f"Limite: {hora_limite.strftime('%H:%M')}"
                    )
                )

                st.success("Configuração salva com sucesso.")
                st.rerun()

            except Exception as erro:
                st.error(
                    f"Erro ao salvar configuração: {erro}"
                )


def aba_envio_manual(supabase):
    st.subheader("📨 Envio manual do dia")

    st.caption(
        "Abra a conversa, envie a mensagem no WhatsApp e depois "
        "marque o permissionário como enviado."
    )

    config = carregar_config(supabase)
    base_url = str(config.get("base_url") or "").strip()

    if not base_url:
        st.warning(
            "Configure a URL do sistema na aba Mensagem e link."
        )
        return

    if "localhost" in base_url.lower():
        st.warning(
            "A URL ainda aponta para localhost. Permissionários "
            "não conseguirão abrir esse endereço fora do seu computador."
        )

    df_perm = carregar_permissionarios(supabase)

    if df_perm.empty:
        st.info("Nenhum permissionário cadastrado.")
        return

    df_ativos = df_perm[df_perm["ativo"] == True].copy()

    if df_ativos.empty:
        st.info("Nenhum permissionário ativo.")
        return

    disparos_hoje = carregar_disparos_do_dia(supabase)

    df_ativos["enviado"] = df_ativos["id"].apply(
        lambda permissionario_id: (
            disparos_hoje
            .get(int(permissionario_id), {})
            .get("status") == "enviado"
        )
    )

    total_ativos = len(df_ativos)
    total_enviados = int(df_ativos["enviado"].sum())
    total_pendentes = total_ativos - total_enviados

    resumo1, resumo2, resumo3 = st.columns(3)
    resumo1.metric("Ativos", total_ativos)
    resumo2.metric("Enviados hoje", total_enviados)
    resumo3.metric("Pendentes", total_pendentes)

    filtro1, filtro2 = st.columns([3, 1])

    with filtro1:
        busca = st.text_input(
            "Buscar permissionário",
            placeholder="Digite o nome ou telefone",
            key="busca_envio_manual"
        )

    with filtro2:
        somente_pendentes = st.toggle(
            "Somente pendentes",
            value=True,
            key="somente_pendentes_manual"
        )

    df_lista = df_ativos.copy()

    if busca.strip():
        texto_busca = busca.strip().lower()

        filtro_nome = (
            df_lista["nome"]
            .astype(str)
            .str.lower()
            .str.contains(texto_busca, na=False, regex=False)
        )

        filtro_telefone = (
            df_lista["whatsapp"]
            .astype(str)
            .str.contains(texto_busca, na=False, regex=False)
        )

        df_lista = df_lista[filtro_nome | filtro_telefone]

    if somente_pendentes:
        df_lista = df_lista[df_lista["enviado"] == False]

    df_lista = df_lista.sort_values("nome")

    if df_lista.empty:
        st.success(
            "Nenhum envio pendente com os filtros selecionados."
        )
        return

    col_itens, col_pagina = st.columns(2)

    with col_itens:
        itens_por_pagina = st.selectbox(
            "Itens por página",
            [10, 20, 50],
            index=1,
            key="itens_pagina_envio_manual"
        )

    total_resultados = len(df_lista)
    total_paginas = max(
        1,
        (total_resultados + itens_por_pagina - 1)
        // itens_por_pagina
    )

    pagina_key = "pagina_envio_manual"
    if pagina_key not in st.session_state:
        st.session_state[pagina_key] = 1

    if st.session_state[pagina_key] > total_paginas:
        st.session_state[pagina_key] = 1

    with col_pagina:
        pagina_atual = st.number_input(
            "Página",
            min_value=1,
            max_value=total_paginas,
            step=1,
            key=pagina_key
        )

    inicio = (int(pagina_atual) - 1) * itens_por_pagina
    fim = inicio + itens_por_pagina
    df_pagina = df_lista.iloc[inicio:fim]

    st.caption(
        f"Mostrando {inicio + 1} a "
        f"{min(fim, total_resultados)} de {total_resultados}"
    )

    for _, linha in df_pagina.iterrows():
        permissionario_id = int(linha["id"])
        nome = str(linha.get("nome", "")).strip()
        whatsapp = normalizar_whatsapp(
            linha.get("whatsapp", "")
        )
        enviado = bool(linha.get("enviado", False))

        if not whatsapp:
            st.warning(
                f"{nome} está sem um WhatsApp válido."
            )
            continue

        try:
            link, valido_ate = gerar_ou_atualizar_link(
                supabase=supabase,
                permissionario_id=permissionario_id,
                config=config
            )

            texto = montar_mensagem_whatsapp(
                nome,
                link,
                config
            )

            wa_link = (
                f"https://wa.me/{whatsapp}"
                f"?text={quote(texto, safe='')}"
            )

            with st.container(border=True):
                col_nome, col_status, col_abrir, col_marcar = (
                    st.columns([3, 1.3, 1.8, 1.8])
                )

                with col_nome:
                    st.markdown(f"**{nome}**")
                    st.caption(
                        f"WhatsApp: {whatsapp} | "
                        f"Link válido até {valido_ate.strftime('%H:%M')}"
                    )

                with col_status:
                    if enviado:
                        st.markdown("✅ **Enviado**")
                    else:
                        st.markdown("⏳ **Pendente**")

                with col_abrir:
                    st.link_button(
                        "Abrir WhatsApp",
                        wa_link,
                        use_container_width=True
                    )

                with col_marcar:
                    if enviado:
                        st.button(
                            "Enviado",
                            key=f"manual_enviado_{permissionario_id}",
                            disabled=True,
                            use_container_width=True
                        )
                    else:
                        if st.button(
                            "Marcar enviado",
                            key=f"marcar_manual_{permissionario_id}",
                            use_container_width=True
                        ):
                            registrar_disparo(
                                supabase=supabase,
                                data_cotacao=hoje_brasil().isoformat(),
                                permissionario_id=permissionario_id,
                                permissionario_nome=nome,
                                whatsapp=whatsapp,
                                link=link,
                                status="enviado",
                                message_id="manual",
                                erro=None
                            )

                            st.success(
                                f"Envio para {nome} marcado como concluído."
                            )
                            st.rerun()

        except Exception as erro:
            st.error(
                f"Erro ao preparar o envio para {nome}: {erro}"
            )


def aba_respostas_admin(supabase):
    st.subheader("📊 Respostas recebidas")

    data_ref = st.date_input(
        "Data das respostas",
        value=hoje_brasil(),
        key="data_respostas_perm"
    )

    data_str = data_ref.strftime("%Y-%m-%d")
    df_resp = carregar_respostas_permissionarios(
        supabase,
        data_str
    )

    if df_resp.empty:
        st.info("Nenhuma resposta registrada para essa data.")
        return

    df_mostrar = df_resp.copy()
    df_mostrar["preco"] = df_mostrar["preco"].apply(
        lambda valor: (
            f"{float(valor):.2f}".replace(".", ",")
            if pd.notnull(valor)
            else ""
        )
    )

    st.dataframe(
        df_mostrar,
        use_container_width=True,
        hide_index=True
    )



def montar_lista_permissionarios(supabase):
    df_permissionarios = carregar_permissionarios(supabase)

    if df_permissionarios.empty:
        return pd.DataFrame(
            columns=[
                "Nome",
                "WhatsApp",
                "Produtos vinculados"
            ]
        )

    df_permissionarios = df_permissionarios.copy()
    df_permissionarios["id"] = pd.to_numeric(
        df_permissionarios["id"],
        errors="coerce"
    )
    df_permissionarios = df_permissionarios.dropna(subset=["id"])
    df_permissionarios["id"] = df_permissionarios["id"].astype(int)

    resposta = (
        supabase
        .table("permissionario_produtos")
        .select("permissionario_id, produto, ativo")
        .eq("ativo", True)
        .execute()
    )

    df_produtos = pd.DataFrame(resposta.data or [])

    if df_produtos.empty:
        df_lista = df_permissionarios.copy()
        df_lista["produtos_vinculados"] = ""
    else:
        df_produtos["permissionario_id"] = pd.to_numeric(
            df_produtos["permissionario_id"],
            errors="coerce"
        )
        df_produtos = df_produtos.dropna(
            subset=["permissionario_id"]
        )
        df_produtos["permissionario_id"] = (
            df_produtos["permissionario_id"].astype(int)
        )
        df_produtos["produto"] = (
            df_produtos["produto"]
            .astype(str)
            .str.strip()
            .str.upper()
        )

        produtos_agrupados = (
            df_produtos
            .groupby("permissionario_id")["produto"]
            .apply(
                lambda valores: ", ".join(
                    sorted(
                        {
                            produto
                            for produto in valores
                            if produto
                        }
                    )
                )
            )
            .reset_index(name="produtos_vinculados")
        )

        df_lista = df_permissionarios.merge(
            produtos_agrupados,
            left_on="id",
            right_on="permissionario_id",
            how="left"
        )

    df_lista["produtos_vinculados"] = (
        df_lista["produtos_vinculados"]
        .fillna("")
    )

    df_lista = df_lista[
        ["nome", "whatsapp", "produtos_vinculados"]
    ].rename(columns={
        "nome": "Nome",
        "whatsapp": "WhatsApp",
        "produtos_vinculados": "Produtos vinculados"
    })

    return df_lista.sort_values("Nome").reset_index(drop=True)


def gerar_excel_permissionarios(df_lista):
    arquivo = BytesIO()

    with pd.ExcelWriter(
        arquivo,
        engine="openpyxl"
    ) as writer:
        df_lista.to_excel(
            writer,
            index=False,
            sheet_name="Permissionários"
        )

        planilha = writer.sheets["Permissionários"]
        planilha.freeze_panes = "A2"
        planilha.auto_filter.ref = planilha.dimensions

        preenchimento = PatternFill(
            fill_type="solid",
            fgColor="1F4E79"
        )

        for celula in planilha[1]:
            celula.font = Font(
                bold=True,
                color="FFFFFF"
            )
            celula.fill = preenchimento
            celula.alignment = Alignment(
                horizontal="center",
                vertical="center"
            )

        planilha.column_dimensions["A"].width = 32
        planilha.column_dimensions["B"].width = 22
        planilha.column_dimensions["C"].width = 85

        for linha in planilha.iter_rows(min_row=2):
            linha[1].number_format = "@"

            for celula in linha:
                celula.alignment = Alignment(
                    vertical="top",
                    wrap_text=True
                )

    arquivo.seek(0)
    return arquivo.getvalue()


def gerar_pdf_permissionarios(df_lista):
    arquivo = BytesIO()

    documento = SimpleDocTemplate(
        arquivo,
        pagesize=landscape(A4),
        leftMargin=24,
        rightMargin=24,
        topMargin=24,
        bottomMargin=24
    )

    estilos = getSampleStyleSheet()

    estilo_titulo = ParagraphStyle(
        "titulo_lista_permissionarios",
        parent=estilos["Title"],
        fontSize=17,
        leading=20,
        alignment=TA_CENTER,
        spaceAfter=5
    )

    estilo_info = ParagraphStyle(
        "info_lista_permissionarios",
        parent=estilos["Normal"],
        fontSize=8.5,
        leading=11,
        alignment=TA_CENTER,
        textColor=colors.grey
    )

    estilo_cabecalho = ParagraphStyle(
        "cabecalho_lista_permissionarios",
        parent=estilos["Normal"],
        fontName="Helvetica-Bold",
        fontSize=9,
        leading=11,
        alignment=TA_CENTER,
        textColor=colors.white
    )

    estilo_celula = ParagraphStyle(
        "celula_lista_permissionarios",
        parent=estilos["Normal"],
        fontSize=8.5,
        leading=11,
        alignment=TA_LEFT
    )

    elementos = [
        Paragraph("Lista de Permissionários", estilo_titulo),
        Paragraph(
            "Nome, WhatsApp e produtos vinculados",
            estilo_info
        ),
        Paragraph(
            f"Gerado em {agora_brasil().strftime('%d/%m/%Y %H:%M')}",
            estilo_info
        ),
        Spacer(1, 12)
    ]

    dados = [[
        Paragraph("Nome", estilo_cabecalho),
        Paragraph("WhatsApp", estilo_cabecalho),
        Paragraph("Produtos vinculados", estilo_cabecalho)
    ]]

    for _, linha in df_lista.iterrows():
        dados.append([
            Paragraph(
                escape(str(linha.get("Nome", ""))),
                estilo_celula
            ),
            Paragraph(
                escape(str(linha.get("WhatsApp", ""))),
                estilo_celula
            ),
            Paragraph(
                escape(
                    str(linha.get("Produtos vinculados", ""))
                    or "Sem produtos vinculados"
                ),
                estilo_celula
            )
        ])

    tabela = Table(
        dados,
        colWidths=[190, 135, 425],
        repeatRows=1
    )

    tabela.setStyle(TableStyle([
        (
            "BACKGROUND",
            (0, 0),
            (-1, 0),
            colors.HexColor("#1F4E79")
        ),
        (
            "TEXTCOLOR",
            (0, 0),
            (-1, 0),
            colors.white
        ),
        (
            "GRID",
            (0, 0),
            (-1, -1),
            0.35,
            colors.grey
        ),
        (
            "VALIGN",
            (0, 0),
            (-1, -1),
            "TOP"
        ),
        (
            "ROWBACKGROUNDS",
            (0, 1),
            (-1, -1),
            [colors.white, colors.whitesmoke]
        ),
        (
            "LEFTPADDING",
            (0, 0),
            (-1, -1),
            5
        ),
        (
            "RIGHTPADDING",
            (0, 0),
            (-1, -1),
            5
        ),
        (
            "TOPPADDING",
            (0, 0),
            (-1, -1),
            5
        ),
        (
            "BOTTOMPADDING",
            (0, 0),
            (-1, -1),
            5
        )
    ]))

    elementos.append(tabela)
    documento.build(elementos)

    arquivo.seek(0)
    return arquivo.getvalue()


def aba_lista_permissionarios(supabase):
    st.subheader("📋 Lista de permissionários")

    try:
        df_lista = montar_lista_permissionarios(supabase)
    except Exception as erro:
        st.error(
            f"Erro ao montar a lista de permissionários: {erro}"
        )
        return

    if df_lista.empty:
        st.info("Nenhum permissionário cadastrado.")
        return

    st.caption(
        f"Total de permissionários: {len(df_lista)}"
    )

    st.dataframe(
        df_lista,
        width="stretch",
        hide_index=True
    )

    try:
        arquivo_excel = gerar_excel_permissionarios(df_lista)
        arquivo_pdf = gerar_pdf_permissionarios(df_lista)
    except Exception as erro:
        st.error(
            f"Erro ao preparar os arquivos para download: {erro}"
        )
        return

    coluna_excel, coluna_pdf = st.columns(2)

    with coluna_excel:
        st.download_button(
            "📥 Baixar Excel",
            data=arquivo_excel,
            file_name="lista_permissionarios.xlsx",
            mime=(
                "application/vnd.openxmlformats-officedocument."
                "spreadsheetml.sheet"
            ),
            width="stretch"
        )

    with coluna_pdf:
        st.download_button(
            "📥 Baixar PDF",
            data=arquivo_pdf,
            file_name="lista_permissionarios.pdf",
            mime="application/pdf",
            width="stretch"
        )


def tela_envio_links_permissionarios(supabase):
    st.title("📨 Envio de links aos permissionários")
    aba_envio_manual(supabase)


def tela_permissionarios_cotacao(supabase):
    st.title("🧑‍🌾 Permissionários")

    abas = st.tabs([
        "Envio de links",
        "Respostas"
    ])

    with abas[0]:
        aba_envio_manual(supabase)

    with abas[1]:
        tela_respostas_permissionarios(
            supabase,
            mostrar_titulo=False
        )

def tela_permissionarios_admin(
    supabase,
    carregar_produtos,
    corrigir_classe,
    registrar_acao
):
    st.title("🧑‍🌾 Permissionários")

    abas = st.tabs([
        "Cadastro",
        "Editar / Excluir",
        "Produtos vinculados",
        "Mensagem e link",
        "Envio manual",
        "Respostas",
        "Lista / Exportação"
    ])

    with abas[0]:
        aba_cadastro_permissionario(
            supabase,
            registrar_acao
        )

    with abas[1]:
        aba_editar_excluir_permissionario(
            supabase,
            registrar_acao
        )

    with abas[2]:
        aba_produtos_permissionario(
            supabase,
            carregar_produtos,
            corrigir_classe,
            registrar_acao
        )

    with abas[3]:
        aba_mensagem_link(
            supabase,
            registrar_acao
        )

    with abas[4]:
        aba_envio_manual(supabase)

    with abas[5]:
        aba_respostas_admin(supabase)

    with abas[6]:
        aba_lista_permissionarios(supabase)



def tela_respostas_permissionarios(supabase, mostrar_titulo=True):
    if mostrar_titulo:
        st.title("📊 Respostas dos Permissionários")

    # ================= FILTROS =================
    st.subheader("🔎 Filtros")

    col_data, col_perm, col_prod = st.columns(3)

    with col_data:
        data_ref = st.date_input(
            "Data das respostas",
            value=hoje_brasil(),
            key="data_respostas_permissionarios_geral"
        )

    data_str = data_ref.strftime("%Y-%m-%d")

    # ================= CARREGAR DADOS =================
    df_resp = carregar_respostas_permissionarios(supabase, data_str)
    df_fotos = carregar_fotos_permissionarios(supabase, data_str)

    # Monta lista de permissionários
    permissionarios_lista = []

    if not df_resp.empty and "permissionario_nome" in df_resp.columns:
        permissionarios_lista += df_resp["permissionario_nome"].dropna().astype(str).tolist()

    if not df_fotos.empty and "permissionario_nome" in df_fotos.columns:
        permissionarios_lista += df_fotos["permissionario_nome"].dropna().astype(str).tolist()

    permissionarios_lista = sorted(list(set(permissionarios_lista)))

    # Monta lista de produtos
    produtos_lista = []

    if not df_resp.empty:
        produtos_lista += df_resp["produto"].dropna().astype(str).str.upper().tolist()

    if not df_fotos.empty:
        produtos_lista += df_fotos["produto"].dropna().astype(str).str.upper().tolist()

    produtos_lista = sorted(list(set(produtos_lista)))

    with col_perm:
        filtro_permissionario = st.selectbox(
            "Permissionário",
            ["Todos"] + permissionarios_lista,
            key="filtro_resp_permissionario"
        )

    with col_prod:
        filtro_produto = st.selectbox(
            "Produto",
            ["Todos"] + produtos_lista,
            key="filtro_resp_produto"
        )

    if df_resp.empty and df_fotos.empty:
        st.warning(
            f"Nenhuma resposta registrada para {data_ref.strftime('%d/%m/%Y')}."
        )
        return

    # ================= PREÇOS =================
    st.subheader("💰 Preços enviados")

    if df_resp.empty:
        st.info("Nenhum preço enviado para essa data.")
    else:
        df_precos = df_resp.copy()

        df_precos["produto"] = df_precos["produto"].astype(str).str.strip().str.upper()

        if filtro_permissionario != "Todos":
            df_precos = df_precos[
                df_precos["permissionario_nome"] == filtro_permissionario
            ]

        if filtro_produto != "Todos":
            df_precos = df_precos[
                df_precos["produto"] == filtro_produto
            ]

        if df_precos.empty:
            st.warning("Nenhum preço encontrado com os filtros selecionados.")
        else:
            if "numero_preco" in df_precos.columns:
                df_precos = df_precos.sort_values(
                    ["permissionario_nome", "produto", "numero_preco"]
                )
            else:
                df_precos = df_precos.sort_values(
                    ["permissionario_nome", "produto"]
                )

            df_mostrar = df_precos.copy()

            if "preco" in df_mostrar.columns:
                df_mostrar["preco"] = df_mostrar["preco"].apply(
                    lambda x: f"{float(x):.2f}".replace(".", ",") if pd.notnull(x) else ""
                )

            if "enviado_em" in df_mostrar.columns:
                df_mostrar["enviado_em"] = pd.to_datetime(
                    df_mostrar["enviado_em"],
                    errors="coerce"
                ).dt.strftime("%d/%m/%Y %H:%M")

            colunas_exibir = [
                "permissionario_nome",
                "produto",
                "classe",
                "numero_preco",
                "preco",
                "enviado_em"
            ]

            colunas_existentes = [c for c in colunas_exibir if c in df_mostrar.columns]

            st.dataframe(
                df_mostrar[colunas_existentes],
                use_container_width=True
            )

    # ================= FOTOS =================
    st.subheader("📷 Fotos enviadas")

    if df_fotos.empty:
        st.info("Nenhuma foto enviada para essa data.")
    else:
        df_img = df_fotos.copy()

        df_img["produto"] = df_img["produto"].astype(str).str.strip().str.upper()

        if filtro_permissionario != "Todos":
            df_img = df_img[
                df_img["permissionario_nome"] == filtro_permissionario
            ]

        if filtro_produto != "Todos":
            df_img = df_img[
                df_img["produto"] == filtro_produto
            ]

        if df_img.empty:
            st.warning("Nenhuma foto encontrada com os filtros selecionados.")
        else:
            for _, row in df_img.iterrows():
                with st.container():
                    st.markdown(f"### 📦 {row.get('produto', '')}")
                    st.caption(
                        f"Permissionário: {row.get('permissionario_nome', '')} | "
                        f"Classe: {row.get('classe', '')}"
                    )

                    st.image(
                        row.get("foto_url", ""),
                        width=350
                    )

                    st.markdown(
                        f"[Abrir foto em nova aba]({row.get('foto_url', '')})"
                    )

                    st.divider()